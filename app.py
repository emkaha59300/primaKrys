from flask import Flask, request, render_template, send_file
import requests
import pandas as pd
from bs4 import BeautifulSoup
import os
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        url = request.form.get("url")

        try:
            # Pobranie zawartości strony
            response = requests.get(url)
            soup = BeautifulSoup(response.text, "html.parser")

            # Pobranie danych z <span class="nms">
            nms_data = [span.text.strip() for span in soup.find_all("span", class_="nms")]

            # Pobranie danych z <span class="res rsl"> lub <span class="res rsl postp">
            res_rsl_data = [
                span.text.strip() 
                for span in soup.find_all("span", class_=lambda x: x and ("res rsl" in x or "res rsl postp" in x))
            ]
            # Dopasowanie długości list (jeśli różna liczba elementów)
            max_length = max(len(nms_data), len(res_rsl_data))
            nms_data.extend([""] * (max_length - len(nms_data)))
            res_rsl_data.extend([""] * (max_length - len(res_rsl_data)))

            # Tworzenie DataFrame
            today = datetime.now().strftime("%Y-%m-%d")

            df = pd.DataFrame({"MECZ": nms_data, "WYNIK": res_rsl_data})

            # Generowanie nazwy pliku z datą
            today = datetime.now().strftime("%Y-%m-%d")
            file_name = f"dane_{today}.xlsx"
            file_path = os.path.join(file_name)

            # Zapis do Excela (bez daty)
            df.to_excel(file_path, index=False)

            # Dodanie daty w pierwszym wierszu
            wb = load_workbook(file_path)
            ws = wb.active
            ws.insert_rows(1)  # Wstawienie nowego pierwszego wiersza
            ws["A1"] = f"Data generacji: {today}"  # Wpisanie daty w A1
            wb.save(file_path)

            return send_file(file_path, as_attachment=True)

        except Exception as e:
            return f"Błąd: {str(e)}"

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
